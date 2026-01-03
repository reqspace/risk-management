"""
Process module for extracting risks, tasks, and decisions from content using Claude API.
Updates Risk Register Excel files with extracted data.
"""

import json
import os
from datetime import datetime
from pathlib import Path
from typing import Optional

import anthropic
import openpyxl
from dotenv import load_dotenv

# Load environment variables
load_dotenv()


def extract_data_with_claude(content: str, project_code: str, source_type: str) -> dict:
    """
    Send content to Claude API to extract risks, tasks, and decisions.
    Returns structured data for Excel update.
    """
    client = anthropic.Anthropic()

    extraction_prompt = f"""Analyze this {source_type} content for project {project_code} and extract the following information in JSON format:

1. **Risks**: Any mentioned risks, concerns, potential issues, or problems that could affect the project.
2. **Tasks**: Action items, assignments, or work that needs to be done.
3. **Decisions**: Any decisions made or conclusions reached.

For each RISK, provide:
- title: Brief title (max 100 chars)
- description: Detailed description
- category: One of [Technical, Schedule, Budget, Resource, External, Safety, Quality]
- probability: One of [Low, Medium, High]
- impact: One of [Low, Medium, High]
- owner: Person responsible (if mentioned, otherwise "TBD")
- mitigation_plan: Any mentioned mitigation strategies (if any, otherwise empty string)

For each TASK, provide:
- task: Description of the task
- owner: Person assigned (if mentioned, otherwise "TBD")
- due_date: Due date if mentioned (format: YYYY-MM-DD), otherwise null

For each DECISION, provide:
- decision: What was decided
- context: Why it was decided (if mentioned)

Return ONLY valid JSON in this exact format:
{{
    "risks": [
        {{
            "title": "...",
            "description": "...",
            "category": "...",
            "probability": "...",
            "impact": "...",
            "owner": "...",
            "mitigation_plan": "..."
        }}
    ],
    "tasks": [
        {{
            "task": "...",
            "owner": "...",
            "due_date": null
        }}
    ],
    "decisions": [
        {{
            "decision": "...",
            "context": "..."
        }}
    ]
}}

If no items are found for a category, return an empty array.

CONTENT:
{content}
"""

    message = client.messages.create(
        model="claude-sonnet-4-20250514",
        max_tokens=4096,
        messages=[
            {"role": "user", "content": extraction_prompt}
        ]
    )

    response_text = message.content[0].text

    # Extract JSON from response (handle potential markdown code blocks)
    if "```json" in response_text:
        response_text = response_text.split("```json")[1].split("```")[0]
    elif "```" in response_text:
        response_text = response_text.split("```")[1].split("```")[0]

    return json.loads(response_text.strip())


def calculate_risk_score(probability: str, impact: str) -> int:
    """Calculate risk score from probability and impact."""
    prob_values = {"Low": 1, "Medium": 2, "High": 3}
    impact_values = {"Low": 1, "Medium": 2, "High": 3}

    prob_val = prob_values.get(probability, 2)
    impact_val = impact_values.get(impact, 2)

    return prob_val * impact_val


def get_next_risk_id(worksheet) -> str:
    """Get the next available Risk ID in R-YYYY-NNN format."""
    year = datetime.now().year
    max_num = 0

    for row in range(2, worksheet.max_row + 1):
        cell_value = worksheet.cell(row=row, column=1).value
        if cell_value and str(cell_value).startswith(f"R-{year}-"):
            try:
                num = int(str(cell_value).split("-")[-1])
                max_num = max(max_num, num)
            except ValueError:
                pass

    return f"R-{year}-{str(max_num + 1).zfill(3)}"


def get_next_task_id(worksheet) -> str:
    """Get the next available Task ID in T-NNN format."""
    max_num = 0

    for row in range(2, worksheet.max_row + 1):
        cell_value = worksheet.cell(row=row, column=1).value
        if cell_value and str(cell_value).startswith("T-"):
            try:
                num = int(str(cell_value).split("-")[-1])
                max_num = max(max_num, num)
            except ValueError:
                pass

    return f"T-{str(max_num + 1).zfill(3)}"


def update_risk_register(workbook, risks: list, source_name: str) -> list:
    """Add new risks to the Risk Register sheet. Returns list of changes made."""
    ws = workbook["Risk Register"]
    changes = []
    today = datetime.now().strftime("%Y-%m-%d")

    for risk in risks:
        next_id = get_next_risk_id(ws)
        risk_score = calculate_risk_score(
            risk.get("probability", "Medium"),
            risk.get("impact", "Medium")
        )

        # Find next empty row
        next_row = ws.max_row + 1

        # Columns: Risk ID, Title, Description, Category, Probability, Impact,
        # Risk Score, Status, Trend, Owner, Mitigation Plan, Linked Tasks,
        # Date Identified, Last Updated, Update History, Source, Closed Date, Resolution Notes
        ws.cell(row=next_row, column=1, value=next_id)
        ws.cell(row=next_row, column=2, value=risk.get("title", ""))
        ws.cell(row=next_row, column=3, value=risk.get("description", ""))
        ws.cell(row=next_row, column=4, value=risk.get("category", ""))
        ws.cell(row=next_row, column=5, value=risk.get("probability", "Medium"))
        ws.cell(row=next_row, column=6, value=risk.get("impact", "Medium"))
        ws.cell(row=next_row, column=7, value=risk_score)
        ws.cell(row=next_row, column=8, value="Open")
        ws.cell(row=next_row, column=9, value="New")
        ws.cell(row=next_row, column=10, value=risk.get("owner", "TBD"))
        ws.cell(row=next_row, column=11, value=risk.get("mitigation_plan", ""))
        ws.cell(row=next_row, column=12, value="")  # Linked Tasks
        ws.cell(row=next_row, column=13, value=today)  # Date Identified
        ws.cell(row=next_row, column=14, value=today)  # Last Updated
        ws.cell(row=next_row, column=15, value=f"{today}: Created from {source_name}")
        ws.cell(row=next_row, column=16, value=source_name)

        changes.append(f"Added Risk {next_id}: {risk.get('title', '')}")

    return changes


def update_tasks(workbook, tasks: list, source_name: str) -> list:
    """Add new tasks to the Tasks sheet. Returns list of changes made."""
    ws = workbook["Tasks"]
    changes = []
    today = datetime.now().strftime("%Y-%m-%d")

    for task in tasks:
        next_id = get_next_task_id(ws)

        # Find next empty row
        next_row = ws.max_row + 1

        # Columns: Task ID, Task, Owner, Due Date, Status, Linked Risk,
        # Source, Created Date, Completed Date
        ws.cell(row=next_row, column=1, value=next_id)
        ws.cell(row=next_row, column=2, value=task.get("task", ""))
        ws.cell(row=next_row, column=3, value=task.get("owner", "TBD"))
        ws.cell(row=next_row, column=4, value=task.get("due_date") or "")
        ws.cell(row=next_row, column=5, value="Open")
        ws.cell(row=next_row, column=6, value="")  # Linked Risk
        ws.cell(row=next_row, column=7, value=source_name)
        ws.cell(row=next_row, column=8, value=today)

        changes.append(f"Added Task {next_id}: {task.get('task', '')[:50]}...")

    return changes


def log_update(workbook, source_name: str, source_type: str, changes: list, raw_extract: dict):
    """Add entry to the Update Log sheet."""
    ws = workbook["Update Log"]

    # Find next empty row
    next_row = ws.max_row + 1

    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    changes_text = "\n".join(changes) if changes else "No changes made"
    raw_text = json.dumps(raw_extract, indent=2)

    # Columns: Timestamp, Source, Source Type, Changes Made, Raw Extract
    ws.cell(row=next_row, column=1, value=timestamp)
    ws.cell(row=next_row, column=2, value=source_name)
    ws.cell(row=next_row, column=3, value=source_type)
    ws.cell(row=next_row, column=4, value=changes_text)
    ws.cell(row=next_row, column=5, value=raw_text[:32000])  # Excel cell limit


def process_content(
    project_code: str,
    content: str,
    source_type: str,
    source_name: str,
    base_path: Optional[str] = None
) -> dict:
    """
    Main processing function. Extracts data from content and updates Excel.

    Args:
        project_code: Project code (e.g., "HB")
        content: Text content to process (email body, meeting transcript, etc.)
        source_type: Type of source ("email", "meeting", etc.)
        source_name: Name/identifier of the source
        base_path: Optional base path for Risk Register files

    Returns:
        dict with success status, changes made, and extracted data
    """
    # Determine base path
    if base_path is None:
        base_path = Path(__file__).parent.resolve()
    else:
        base_path = Path(base_path)

    # Determine Risk Register path
    risk_register_path = base_path / "Risk_Registers" / f"Risk_Register_{project_code}.xlsx"

    if not risk_register_path.exists():
        return {
            "success": False,
            "error": f"Risk Register not found at {risk_register_path}",
            "changes": [],
            "extracted_data": None
        }

    # Extract data using Claude API
    try:
        extracted_data = extract_data_with_claude(content, project_code, source_type)
    except anthropic.APIError as e:
        return {
            "success": False,
            "error": f"Claude API error: {str(e)}",
            "changes": [],
            "extracted_data": None
        }
    except json.JSONDecodeError as e:
        return {
            "success": False,
            "error": f"Failed to parse Claude response: {str(e)}",
            "changes": [],
            "extracted_data": None
        }

    # Load Risk Register workbook
    try:
        workbook = openpyxl.load_workbook(risk_register_path)
    except Exception as e:
        return {
            "success": False,
            "error": f"Failed to open Risk Register: {str(e)}",
            "changes": [],
            "extracted_data": extracted_data
        }

    # Update Risk Register with extracted data
    all_changes = []

    if extracted_data.get("risks"):
        risk_changes = update_risk_register(workbook, extracted_data["risks"], source_name)
        all_changes.extend(risk_changes)

    if extracted_data.get("tasks"):
        task_changes = update_tasks(workbook, extracted_data["tasks"], source_name)
        all_changes.extend(task_changes)

    # Log the update
    log_update(workbook, source_name, source_type, all_changes, extracted_data)

    # Save workbook
    try:
        workbook.save(risk_register_path)
    except Exception as e:
        return {
            "success": False,
            "error": f"Failed to save Risk Register: {str(e)}",
            "changes": all_changes,
            "extracted_data": extracted_data
        }

    return {
        "success": True,
        "error": None,
        "changes": all_changes,
        "extracted_data": extracted_data,
        "summary": {
            "risks_added": len(extracted_data.get("risks", [])),
            "tasks_added": len(extracted_data.get("tasks", [])),
            "decisions_found": len(extracted_data.get("decisions", []))
        }
    }


if __name__ == "__main__":
    # CLI usage for testing
    import sys

    if len(sys.argv) < 5:
        print("Usage: python process.py <project_code> <content_file> <source_type> <source_name>")
        print("Example: python process.py HB transcript.txt meeting 'Weekly Standup 2025-01-01'")
        sys.exit(1)

    project = sys.argv[1]
    content_file = sys.argv[2]
    src_type = sys.argv[3]
    src_name = sys.argv[4]

    with open(content_file, 'r', encoding='utf-8') as f:
        content_text = f.read()

    result = process_content(project, content_text, src_type, src_name)

    if result["success"]:
        print("Processing completed successfully!")
        print(f"Summary: {result['summary']}")
        print("Changes made:")
        for change in result["changes"]:
            print(f"  - {change}")
    else:
        print(f"Processing failed: {result['error']}")
        sys.exit(1)
