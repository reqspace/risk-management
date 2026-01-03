"""
Daily Risk Digest Email Generator
Reads Risk Register and generates an HTML email summary.
"""

import os
import smtplib
import io
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
from datetime import datetime, timedelta
from pathlib import Path
from dotenv import load_dotenv
import openpyxl
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
from reportlab.lib.enums import TA_CENTER, TA_LEFT

load_dotenv()

# Email settings
SMTP_SERVER = os.getenv('SMTP_SERVER', 'smtp.office365.com')
SMTP_PORT = int(os.getenv('SMTP_PORT', 587))
EMAIL_FROM = os.getenv('EMAIL_FROM', 'mlaporte@iepwr.com')
EMAIL_TO = os.getenv('EMAIL_TO', 'mlaporte@iepwr.com')
EMAIL_PASSWORD = os.getenv('EMAIL_PASSWORD', '')


def get_risk_register_path(project_code='HB'):
    """Get path to risk register file."""
    base_path = Path(__file__).parent.resolve()
    return base_path / 'Risk_Registers' / f'Risk_Register_{project_code}.xlsx'


def get_all_projects():
    """Get list of all project codes from Risk Register files."""
    base_path = Path(__file__).parent.resolve()
    risk_registers_path = base_path / "Risk_Registers"
    projects = []
    if risk_registers_path.exists():
        for file in risk_registers_path.glob("Risk_Register_*.xlsx"):
            # Skip temp files
            if file.name.startswith('~$'):
                continue
            project_code = file.stem.replace("Risk_Register_", "")
            projects.append(project_code)
    return sorted(projects)


def read_risk_register(project_code='HB'):
    """Read risk register and return risks, tasks, and milestones data."""
    file_path = get_risk_register_path(project_code)

    if not file_path.exists():
        return {'risks': [], 'tasks': [], 'milestones': []}

    wb = openpyxl.load_workbook(file_path)

    # Read risks
    risks = []
    risk_sheet = wb['Risk Register']
    headers = [cell.value for cell in risk_sheet[1]]

    for row_num in range(2, risk_sheet.max_row + 1):
        row_data = {headers[i]: cell.value for i, cell in enumerate(risk_sheet[row_num]) if i < len(headers)}
        if row_data.get('Risk ID'):
            risks.append(row_data)

    # Read tasks
    tasks = []
    task_sheet = wb['Tasks']
    task_headers = [cell.value for cell in task_sheet[1]]

    for row_num in range(2, task_sheet.max_row + 1):
        row_data = {task_headers[i]: cell.value for i, cell in enumerate(task_sheet[row_num]) if i < len(task_headers)}
        if row_data.get('Task ID'):
            tasks.append(row_data)

    # Read milestones
    milestones = []
    if 'Milestones' in wb.sheetnames:
        ms_sheet = wb['Milestones']
        ms_headers = [cell.value for cell in ms_sheet[1]]

        for row_num in range(2, ms_sheet.max_row + 1):
            row_data = {ms_headers[i]: cell.value for i, cell in enumerate(ms_sheet[row_num]) if i < len(ms_headers)}
            if row_data.get('Milestone ID') or row_data.get('Milestone'):
                milestones.append(row_data)

    wb.close()
    return {'risks': risks, 'tasks': tasks, 'milestones': milestones}


def get_digest_data(project_code='HB'):
    """Get all data needed for the digest."""
    data = read_risk_register(project_code)
    risks = data['risks']
    tasks = data['tasks']

    today = datetime.now().date()
    week_from_now = today + timedelta(days=7)
    week_ago = today - timedelta(days=7)

    # Count risks by status
    active_risks = [r for r in risks if r.get('Status') in ['Open', 'Active', 'Escalated']]
    watching_risks = [r for r in risks if r.get('Status') == 'Watching']
    closed_risks = [r for r in risks if r.get('Status') == 'Closed']

    # Count tasks by status
    open_tasks = [t for t in tasks if t.get('Status') in ['Open', 'In Progress', None, '']]

    # Overdue tasks (due date passed and not completed)
    overdue_tasks = []
    for t in open_tasks:
        due_date = t.get('Due Date')
        if due_date:
            if isinstance(due_date, datetime):
                due_date = due_date.date()
            elif isinstance(due_date, str):
                try:
                    due_date = datetime.strptime(due_date, '%Y-%m-%d').date()
                except:
                    continue
            if due_date < today:
                overdue_tasks.append(t)

    # Tasks due this week
    tasks_due_this_week = []
    for t in open_tasks:
        due_date = t.get('Due Date')
        if due_date:
            if isinstance(due_date, datetime):
                due_date = due_date.date()
            elif isinstance(due_date, str):
                try:
                    due_date = datetime.strptime(due_date, '%Y-%m-%d').date()
                except:
                    continue
            if today <= due_date <= week_from_now:
                tasks_due_this_week.append(t)

    # Recently closed risks (last 7 days)
    recently_closed = []
    for r in closed_risks:
        closed_date = r.get('Closed Date')
        if closed_date:
            if isinstance(closed_date, datetime):
                closed_date = closed_date.date()
            elif isinstance(closed_date, str):
                try:
                    closed_date = datetime.strptime(closed_date, '%Y-%m-%d').date()
                except:
                    continue
            if closed_date >= week_ago:
                recently_closed.append(r)

    # Non-closed risks (attention required)
    attention_required = [r for r in risks if r.get('Status') not in ['Closed', None, '']]

    return {
        'active_risks': active_risks,
        'watching_risks': watching_risks,
        'open_tasks': open_tasks,
        'overdue_tasks': overdue_tasks,
        'tasks_due_this_week': tasks_due_this_week,
        'recently_closed': recently_closed,
        'attention_required': attention_required,
        'counts': {
            'active_risks': len(active_risks),
            'watching_risks': len(watching_risks),
            'open_tasks': len(open_tasks),
            'overdue_tasks': len(overdue_tasks)
        }
    }


def get_critical_alerts(projects=None):
    """Get critical alerts across all projects."""
    if projects is None:
        projects = get_all_projects()

    alerts = []

    for project in projects:
        data = read_risk_register(project)

        # Check for critical/high impact risks
        for risk in data['risks']:
            if risk.get('Impact') == 'Critical' or (risk.get('Probability') == 'High' and risk.get('Impact') == 'High'):
                if risk.get('Status') not in ['Closed', 'Mitigated']:
                    alerts.append({
                        'project': project,
                        'type': 'CRITICAL RISK',
                        'id': risk.get('Risk ID'),
                        'title': risk.get('Title'),
                        'description': risk.get('Description'),
                        'status': risk.get('Status'),
                        'owner': risk.get('Owner'),
                        'mitigation': risk.get('Mitigation Plan')
                    })

        # Check for critical/delayed milestones
        for ms in data.get('milestones', []):
            if ms.get('Status') in ['Critical', 'Delayed']:
                alerts.append({
                    'project': project,
                    'type': 'CRITICAL MILESTONE',
                    'id': ms.get('Milestone ID'),
                    'title': ms.get('Milestone'),
                    'description': ms.get('Notes', ''),
                    'status': ms.get('Status'),
                    'baseline': ms.get('Baseline Date'),
                    'current': ms.get('Current Date')
                })

    return alerts


def get_portfolio_summary():
    """Get summary stats for all projects."""
    projects = get_all_projects()
    summary = []

    for project in projects:
        data = read_risk_register(project)
        risks = data['risks']
        tasks = data['tasks']
        milestones = data.get('milestones', [])

        active_risks = len([r for r in risks if r.get('Status') in ['Open', 'Active', 'Escalated']])
        critical_risks = len([r for r in risks if r.get('Impact') == 'Critical' and r.get('Status') != 'Closed'])
        open_tasks = len([t for t in tasks if t.get('Status') not in ['Complete', 'Completed', 'Done']])
        critical_milestones = len([m for m in milestones if m.get('Status') in ['Critical', 'At Risk']])

        # Determine health status
        if critical_risks > 0 or critical_milestones > 0:
            health = 'Critical'
        elif active_risks >= 5:
            health = 'At Risk'
        elif active_risks > 0:
            health = 'Caution'
        else:
            health = 'Healthy'

        summary.append({
            'project': project,
            'health': health,
            'active_risks': active_risks,
            'critical_risks': critical_risks,
            'open_tasks': open_tasks,
            'critical_milestones': critical_milestones
        })

    return summary


def generate_html_email(project_code='HB', include_portfolio=True):
    """Generate HTML email content for the daily digest."""
    data = get_digest_data(project_code)
    today = datetime.now().strftime('%B %d, %Y')

    # Get portfolio summary and critical alerts for multi-project view
    portfolio = get_portfolio_summary() if include_portfolio else []
    critical_alerts = get_critical_alerts() if include_portfolio else []

    # Generate portfolio and critical alerts HTML
    portfolio_html = generate_portfolio_html(portfolio) if include_portfolio and portfolio else ''
    critical_alerts_html = generate_critical_alerts_html(critical_alerts) if critical_alerts else ''

    html = f"""
<!DOCTYPE html>
<html>
<head>
    <style>
        body {{
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            line-height: 1.6;
            color: #333;
            max-width: 900px;
            margin: 0 auto;
            padding: 20px;
        }}
        .header {{
            background: linear-gradient(135deg, #1e3a5f 0%, #2d5a87 100%);
            color: white;
            padding: 25px;
            border-radius: 8px 8px 0 0;
            margin-bottom: 0;
        }}
        .header h1 {{
            margin: 0;
            font-size: 24px;
        }}
        .header .date {{
            opacity: 0.9;
            font-size: 14px;
            margin-top: 5px;
        }}
        .critical-banner {{
            background: #dc3545;
            color: white;
            padding: 15px 20px;
            border-left: 5px solid #a71d2a;
        }}
        .critical-banner h2 {{
            margin: 0 0 10px 0;
            font-size: 16px;
        }}
        .critical-item {{
            background: rgba(255,255,255,0.1);
            padding: 10px 15px;
            margin-bottom: 8px;
            border-radius: 4px;
        }}
        .critical-item .title {{
            font-weight: bold;
        }}
        .critical-item .description {{
            font-size: 13px;
            opacity: 0.9;
            margin-top: 4px;
        }}
        .portfolio-section {{
            background: #e9ecef;
            padding: 20px;
            border: 1px solid #dee2e6;
        }}
        .portfolio-section h2 {{
            color: #1e3a5f;
            font-size: 16px;
            margin-top: 0;
            margin-bottom: 15px;
        }}
        .portfolio-grid {{
            display: flex;
            flex-wrap: wrap;
            gap: 12px;
        }}
        .project-card {{
            background: white;
            border-radius: 6px;
            padding: 12px 16px;
            min-width: 180px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
            border-left: 4px solid #28a745;
        }}
        .project-card.critical {{
            border-left-color: #dc3545;
        }}
        .project-card.at-risk {{
            border-left-color: #ffc107;
        }}
        .project-card.caution {{
            border-left-color: #fd7e14;
        }}
        .project-card .project-name {{
            font-weight: bold;
            font-size: 16px;
            color: #1e3a5f;
        }}
        .project-card .project-status {{
            font-size: 11px;
            font-weight: bold;
            text-transform: uppercase;
            margin-top: 2px;
        }}
        .project-card .project-status.critical {{
            color: #dc3545;
        }}
        .project-card .project-status.at-risk {{
            color: #ffc107;
        }}
        .project-card .project-status.caution {{
            color: #fd7e14;
        }}
        .project-card .project-status.healthy {{
            color: #28a745;
        }}
        .project-card .project-stats {{
            font-size: 12px;
            color: #666;
            margin-top: 6px;
        }}
        .section {{
            background: #f8f9fa;
            border: 1px solid #e9ecef;
            border-top: none;
            padding: 20px;
            margin-bottom: 20px;
        }}
        .section:last-child {{
            border-radius: 0 0 8px 8px;
        }}
        .section h2 {{
            color: #1e3a5f;
            font-size: 18px;
            margin-top: 0;
            padding-bottom: 10px;
            border-bottom: 2px solid #2d5a87;
        }}
        .at-a-glance {{
            display: flex;
            flex-wrap: wrap;
            gap: 15px;
        }}
        .stat-box {{
            background: white;
            border-radius: 8px;
            padding: 15px 20px;
            min-width: 150px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }}
        .stat-box .number {{
            font-size: 28px;
            font-weight: bold;
            color: #1e3a5f;
        }}
        .stat-box .label {{
            color: #666;
            font-size: 12px;
            text-transform: uppercase;
        }}
        .stat-box.warning .number {{
            color: #dc3545;
        }}
        .risk-item, .task-item {{
            background: white;
            border-radius: 6px;
            padding: 12px 15px;
            margin-bottom: 10px;
            border-left: 4px solid #2d5a87;
            box-shadow: 0 1px 3px rgba(0,0,0,0.1);
        }}
        .risk-item.high {{
            border-left-color: #dc3545;
        }}
        .risk-item.medium {{
            border-left-color: #ffc107;
        }}
        .risk-item .title {{
            font-weight: 600;
            color: #1e3a5f;
        }}
        .risk-item .meta, .task-item .meta {{
            font-size: 12px;
            color: #666;
            margin-top: 5px;
        }}
        .closed-item {{
            background: #d4edda;
            border-left-color: #28a745;
        }}
        .owner-group {{
            margin-bottom: 15px;
        }}
        .owner-group h3 {{
            color: #495057;
            font-size: 14px;
            margin-bottom: 8px;
        }}
        .empty-message {{
            color: #666;
            font-style: italic;
            padding: 10px;
        }}
        .footer {{
            text-align: center;
            color: #666;
            font-size: 12px;
            padding: 20px;
        }}
    </style>
</head>
<body>
    <div class="header">
        <h1>LDES Program Daily Risk Digest</h1>
        <div class="date">{today}</div>
    </div>

    {critical_alerts_html}

    {portfolio_html}

    <div class="section">
        <h2>{project_code} Project - At A Glance</h2>
        <div class="at-a-glance">
            <div class="stat-box">
                <div class="number">{data['counts']['active_risks']}</div>
                <div class="label">Active Risks</div>
            </div>
            <div class="stat-box">
                <div class="number">{data['counts']['watching_risks']}</div>
                <div class="label">Watching</div>
            </div>
            <div class="stat-box">
                <div class="number">{data['counts']['open_tasks']}</div>
                <div class="label">Open Tasks</div>
            </div>
            <div class="stat-box {'warning' if data['counts']['overdue_tasks'] > 0 else ''}">
                <div class="number">{data['counts']['overdue_tasks']}</div>
                <div class="label">Overdue Tasks</div>
            </div>
        </div>
    </div>

    <div class="section">
        <h2>Attention Required</h2>
        {generate_attention_required_html(data['attention_required'])}
    </div>

    <div class="section">
        <h2>Tasks Due This Week</h2>
        {generate_tasks_html(data['tasks_due_this_week'])}
    </div>

    <div class="section">
        <h2>Recently Closed (Last 7 Days)</h2>
        {generate_recently_closed_html(data['recently_closed'])}
    </div>

    <div class="footer">
        Generated by LDES Risk Management System<br>
        {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
    </div>
</body>
</html>
"""
    return html


def generate_portfolio_html(portfolio):
    """Generate HTML for portfolio overview section."""
    if not portfolio:
        return ''

    cards_html = ''
    for proj in portfolio:
        health_class = proj['health'].lower().replace(' ', '-')
        cards_html += f"""
        <div class="project-card {health_class}">
            <div class="project-name">{proj['project']}</div>
            <div class="project-status {health_class}">{proj['health']}</div>
            <div class="project-stats">
                {proj['active_risks']} risks | {proj['open_tasks']} tasks
                {f" | {proj['critical_milestones']} critical" if proj['critical_milestones'] > 0 else ""}
            </div>
        </div>
        """

    return f"""
    <div class="portfolio-section">
        <h2>Portfolio Overview</h2>
        <div class="portfolio-grid">
            {cards_html}
        </div>
    </div>
    """


def generate_critical_alerts_html(alerts):
    """Generate HTML for critical alerts banner."""
    if not alerts:
        return ''

    items_html = ''
    for alert in alerts[:5]:  # Limit to top 5 critical alerts
        items_html += f"""
        <div class="critical-item">
            <div class="title">[{alert['project']}] {alert['type']}: {alert['title']}</div>
            <div class="description">{alert.get('description', '') or alert.get('mitigation', '')}</div>
        </div>
        """

    return f"""
    <div class="critical-banner">
        <h2>CRITICAL ALERTS ({len(alerts)})</h2>
        {items_html}
    </div>
    """


def generate_attention_required_html(risks):
    """Generate HTML for attention required section."""
    if not risks:
        return '<p class="empty-message">No active risks requiring attention.</p>'

    html = ''
    for risk in risks:
        probability = risk.get('Probability', 'Unknown')
        priority_class = 'high' if probability == 'High' else 'medium' if probability == 'Medium' else ''

        owner = risk.get('Owner', 'Unassigned')
        last_updated = risk.get('Last Updated', '')
        if isinstance(last_updated, datetime):
            last_updated = last_updated.strftime('%Y-%m-%d')

        html += f"""
        <div class="risk-item {priority_class}">
            <div class="title">{risk.get('Risk ID', '')}: {risk.get('Title', 'Untitled')}</div>
            <div class="meta">
                Owner: {owner} |
                Status: {risk.get('Status', 'Unknown')} |
                Priority: {probability}/{risk.get('Impact', 'Unknown')} |
                Last Updated: {last_updated}
            </div>
        </div>
        """
    return html


def generate_tasks_html(tasks):
    """Generate HTML for tasks section, grouped by owner."""
    if not tasks:
        return '<p class="empty-message">No tasks due this week.</p>'

    # Group by owner
    by_owner = {}
    for task in tasks:
        owner = task.get('Owner', 'Unassigned') or 'Unassigned'
        if owner not in by_owner:
            by_owner[owner] = []
        by_owner[owner].append(task)

    html = ''
    for owner, owner_tasks in sorted(by_owner.items()):
        html += f'<div class="owner-group"><h3>{owner}</h3>'
        for task in owner_tasks:
            due_date = task.get('Due Date', '')
            if isinstance(due_date, datetime):
                due_date = due_date.strftime('%Y-%m-%d')

            html += f"""
            <div class="task-item">
                <div class="title">{task.get('Task ID', '')}: {task.get('Task', 'Untitled')}</div>
                <div class="meta">Due: {due_date or 'No date set'} | Source: {task.get('Source', 'Unknown')}</div>
            </div>
            """
        html += '</div>'
    return html


def generate_recently_closed_html(risks):
    """Generate HTML for recently closed section."""
    if not risks:
        return '<p class="empty-message">No risks closed in the last 7 days.</p>'

    html = ''
    for risk in risks:
        closed_date = risk.get('Closed Date', '')
        if isinstance(closed_date, datetime):
            closed_date = closed_date.strftime('%Y-%m-%d')

        html += f"""
        <div class="risk-item closed-item">
            <div class="title">{risk.get('Risk ID', '')}: {risk.get('Title', 'Untitled')}</div>
            <div class="meta">
                Closed: {closed_date} |
                Resolution: {risk.get('Resolution Notes', 'No notes') or 'No notes'}
            </div>
        </div>
        """
    return html


def generate_pdf_report(project_code='HB'):
    """Generate a PDF report and return as bytes."""
    data = get_digest_data(project_code)
    today = datetime.now().strftime('%B %d, %Y')

    # Create PDF in memory
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)

    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=20,
        textColor=colors.HexColor('#1e3a5f'),
        spaceAfter=6
    )
    subtitle_style = ParagraphStyle(
        'Subtitle',
        parent=styles['Normal'],
        fontSize=12,
        textColor=colors.HexColor('#666666'),
        spaceAfter=20
    )
    section_style = ParagraphStyle(
        'Section',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#1e3a5f'),
        spaceBefore=15,
        spaceAfter=10
    )
    normal_style = ParagraphStyle(
        'CustomNormal',
        parent=styles['Normal'],
        fontSize=10,
        spaceAfter=6
    )

    story = []

    # Header
    story.append(Paragraph(f"Executive Risk Report - {project_code}", title_style))
    story.append(Paragraph(today, subtitle_style))
    story.append(HRFlowable(width="100%", thickness=2, color=colors.HexColor('#1e3a5f')))
    story.append(Spacer(1, 15))

    # At a Glance stats
    story.append(Paragraph("At A Glance", section_style))

    stats_data = [
        ['Active Risks', 'Watching', 'Open Tasks', 'Overdue'],
        [str(data['counts']['active_risks']),
         str(data['counts']['watching_risks']),
         str(data['counts']['open_tasks']),
         str(data['counts']['overdue_tasks'])]
    ]

    stats_table = Table(stats_data, colWidths=[1.5*inch, 1.5*inch, 1.5*inch, 1.5*inch])
    stats_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f8f9fa')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#666666')),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 1), (-1, 1), 18),
        ('TEXTCOLOR', (0, 1), (0, 1), colors.HexColor('#dc3545')),  # Active - red
        ('TEXTCOLOR', (1, 1), (1, 1), colors.HexColor('#ffc107')),  # Watching - yellow
        ('TEXTCOLOR', (2, 1), (2, 1), colors.HexColor('#0d6efd')),  # Open - blue
        ('TEXTCOLOR', (3, 1), (3, 1), colors.HexColor('#dc3545') if data['counts']['overdue_tasks'] > 0 else colors.HexColor('#28a745')),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
        ('TOPPADDING', (0, 0), (-1, -1), 10),
        ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#dee2e6')),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#dee2e6')),
    ]))
    story.append(stats_table)
    story.append(Spacer(1, 15))

    # Attention Required (Active Risks)
    story.append(Paragraph("Attention Required", section_style))

    if data['attention_required']:
        risk_data = [['Risk ID', 'Title', 'Priority', 'Owner', 'Status']]
        for risk in data['attention_required'][:10]:
            risk_data.append([
                risk.get('Risk ID', ''),
                Paragraph(str(risk.get('Title', 'Untitled'))[:50], normal_style),
                f"{risk.get('Probability', '-')}/{risk.get('Impact', '-')}",
                risk.get('Owner', 'Unassigned')[:15],
                risk.get('Status', 'Unknown')
            ])

        risk_table = Table(risk_data, colWidths=[0.7*inch, 2.5*inch, 0.9*inch, 1*inch, 0.8*inch])
        risk_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a5f')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#dee2e6')),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#dee2e6')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        story.append(risk_table)
    else:
        story.append(Paragraph("No active risks requiring attention.", normal_style))

    story.append(Spacer(1, 15))

    # Tasks Due This Week
    story.append(Paragraph("Tasks Due This Week", section_style))

    if data['tasks_due_this_week']:
        task_data = [['Task ID', 'Task', 'Owner', 'Due Date']]
        for task in data['tasks_due_this_week'][:8]:
            due_date = task.get('Due Date', '')
            if isinstance(due_date, datetime):
                due_date = due_date.strftime('%Y-%m-%d')
            task_data.append([
                task.get('Task ID', ''),
                Paragraph(str(task.get('Task', 'Untitled'))[:60], normal_style),
                task.get('Owner', 'Unassigned')[:15],
                str(due_date)
            ])

        task_table = Table(task_data, colWidths=[0.7*inch, 3.3*inch, 1*inch, 0.9*inch])
        task_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0d6efd')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#dee2e6')),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#dee2e6')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        story.append(task_table)
    else:
        story.append(Paragraph("No tasks due this week.", normal_style))

    story.append(Spacer(1, 15))

    # Recently Closed
    if data['recently_closed']:
        story.append(Paragraph("Recently Closed (Last 7 Days)", section_style))
        for risk in data['recently_closed'][:5]:
            story.append(Paragraph(
                f"<b>{risk.get('Risk ID', '')}</b>: {risk.get('Title', 'Untitled')}",
                normal_style
            ))

    # Footer
    story.append(Spacer(1, 30))
    story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor('#dee2e6')))
    footer_style = ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, textColor=colors.gray, alignment=TA_CENTER)
    story.append(Paragraph(f"Generated by Risk Management System | {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", footer_style))

    # Build PDF
    doc.build(story)
    buffer.seek(0)
    return buffer.getvalue()


def send_digest_email(project_code='HB', recipients=None, attach_pdf=True):
    """Generate and send the daily digest email with optional PDF attachment."""
    if not EMAIL_PASSWORD:
        print("[Daily Digest] ERROR: EMAIL_PASSWORD not set in .env")
        return False

    html_content = generate_html_email(project_code)
    today = datetime.now().strftime('%Y-%m-%d')

    # Handle multiple recipients
    if recipients is None:
        recipients = EMAIL_TO

    # Convert comma-separated string to list
    if isinstance(recipients, str):
        recipient_list = [r.strip() for r in recipients.split(',')]
    else:
        recipient_list = recipients

    # Create message
    msg = MIMEMultipart('mixed')
    msg['Subject'] = f'Daily Risk Digest - {project_code} Project - {today}'
    msg['From'] = EMAIL_FROM
    msg['To'] = ', '.join(recipient_list)

    # Create alternative part for HTML content
    msg_alternative = MIMEMultipart('alternative')
    html_part = MIMEText(html_content, 'html')
    msg_alternative.attach(html_part)
    msg.attach(msg_alternative)

    # Attach PDF if requested
    if attach_pdf:
        try:
            pdf_bytes = generate_pdf_report(project_code)
            pdf_attachment = MIMEApplication(pdf_bytes, _subtype='pdf')
            pdf_filename = f'Risk_Report_{project_code}_{today}.pdf'
            pdf_attachment.add_header('Content-Disposition', 'attachment', filename=pdf_filename)
            msg.attach(pdf_attachment)
            print(f"[Daily Digest] PDF attached: {pdf_filename}")
        except Exception as e:
            print(f"[Daily Digest] Warning: Could not generate PDF attachment: {e}")

    try:
        # Connect and send
        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
            server.starttls()
            server.login(EMAIL_FROM, EMAIL_PASSWORD)
            server.sendmail(EMAIL_FROM, recipient_list, msg.as_string())

        print(f"[Daily Digest] Email sent successfully to {', '.join(recipient_list)}")
        return True

    except smtplib.SMTPAuthenticationError as e:
        print(f"[Daily Digest] Authentication failed: {e}")
        return False
    except smtplib.SMTPException as e:
        print(f"[Daily Digest] SMTP error: {e}")
        return False
    except Exception as e:
        print(f"[Daily Digest] Error sending email: {e}")
        return False


def preview_digest(project_code='HB'):
    """Generate and save a preview of the digest email."""
    html_content = generate_html_email(project_code)
    preview_path = Path(__file__).parent / 'digest_preview.html'

    with open(preview_path, 'w') as f:
        f.write(html_content)

    print(f"[Daily Digest] Preview saved to {preview_path}")
    return preview_path


if __name__ == '__main__':
    import sys

    if len(sys.argv) > 1 and sys.argv[1] == '--send':
        send_digest_email('HB')
    else:
        # Preview mode
        preview_path = preview_digest('HB')
        print(f"Open {preview_path} in a browser to preview the email")

        # Also print summary
        data = get_digest_data('HB')
        print(f"\nDigest Summary:")
        print(f"  Active Risks: {data['counts']['active_risks']}")
        print(f"  Watching: {data['counts']['watching_risks']}")
        print(f"  Open Tasks: {data['counts']['open_tasks']}")
        print(f"  Overdue Tasks: {data['counts']['overdue_tasks']}")
        print(f"  Tasks Due This Week: {len(data['tasks_due_this_week'])}")
        print(f"  Recently Closed: {len(data['recently_closed'])}")
