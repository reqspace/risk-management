#!/usr/bin/env python3
"""
Process meeting transcripts using Claude API to extract risks, tasks, and decisions.
Updates the project Risk Register Excel file and logs all changes.
"""

import argparse
import json
import os
import sys
from datetime import datetime
from pathlib import Path

import anthropic
import openpyxl
from openpyxl.utils import get_column_letter


def parse_arguments():
    """Parse command line arguments."""
    parser = argparse.ArgumentParser(
        description="Process meeting transcripts to extract risks, tasks, and decisions."
    )
    parser.add_argument(
        "project_code",
        type=str,
        help="Project code (e.g., HB, NSD, RH)"
    )
    parser.add_argument(
        "transcript_path",
        type=str,
        help="Path to the transcript file"
    )
    parser.add_argument(
        "--risk-register",
        type=str,
        default=None,
        help="Path to Risk Register Excel file (default: Risk_Registers/Risk_Register_{PROJECT_CODE}.xlsx)"
    )
    return parser.parse_args()


def load_transcript(transcript_path: str) -> str:
    """Load transcript content from file."""
    path = Path(transcript_path)
    if not path.exists():
        raise FileNotFoundError(f"Transcript file not found: {transcript_path}")

    with open(path, 'r', encoding='utf-8') as f:
        return f.read()


def extract_data_with_claude(transcript: str, project_code: str) -> dict:
    """
    Send transcript to Claude API to extract risks, tasks, and decisions.
    Returns structured data for Excel update.
    """
    client = anthropic.Anthropic()

    extraction_prompt = f"""Analyze this meeting transcript for project {project_code} and extract the following information in JSON format:

1. **Risks**: Any mentioned risks, concerns, potential issues, or problems that could affect the project.
2. **Tasks**: Action items, assignments, or work that needs to be done.
3. **Decisions**: Any decisions made during the meeting.

For each RISK, provide:
- title: Brief title (max 100 chars)
- description: Detailed description
- category: One of [Technical, Schedule, Budget, Resource, External, Safety, Quality]
- probability: One of [Low, Medium, High]
- impact: One of [Low, Medium, High]
- owner: Person responsible (if mentioned, otherwise "TBD")
- mitigation_plan: Any mentioned mitigation strategies (if any)

For each TASK, provide:
- task: Description of the task
- owner: Person assigned (if mentioned, otherwise "TBD")
- due_date: Due date if mentioned (format: YYYY-MM-DD), otherwise null
- linked_risk: Related risk title if applicable, otherwise null

For each DECISION, provide:
- decision: What was decided
- context: Why it was decided (if mentioned)

Return ONLY valid JSON in this exact format:
{{
    "risks": [
        {{
            "title": "...",
            "description": "...",
            "category": "...",
            "probability": "...",
            "impact": "...",
            "owner": "...",
            "mitigation_plan": "..."
        }}
    ],
    "tasks": [
        {{
            "task": "...",
            "owner": "...",
            "due_date": null,
            "linked_risk": null
        }}
    ],
    "decisions": [
        {{
            "decision": "...",
            "context": "..."
        }}
    ]
}}

If no items are found for a category, return an empty array.

TRANSCRIPT:
{transcript}
"""

    message = client.messages.create(
        model="claude-sonnet-4-20250514",
        max_tokens=4096,
        messages=[
            {"role": "user", "content": extraction_prompt}
        ]
    )

    response_text = message.content[0].text

    # Extract JSON from response (handle potential markdown code blocks)
    if "```json" in response_text:
        response_text = response_text.split("```json")[1].split("```")[0]
    elif "```" in response_text:
        response_text = response_text.split("```")[1].split("```")[0]

    return json.loads(response_text.strip())


def calculate_risk_score(probability: str, impact: str) -> int:
    """Calculate risk score from probability and impact."""
    prob_values = {"Low": 1, "Medium": 2, "High": 3}
    impact_values = {"Low": 1, "Medium": 2, "High": 3}

    prob_val = prob_values.get(probability, 2)
    impact_val = impact_values.get(impact, 2)

    return prob_val * impact_val


def get_next_id(worksheet, id_column: int = 1) -> str:
    """Get the next available ID for a worksheet."""
    max_id = 0
    prefix = ""

    for row in range(2, worksheet.max_row + 1):
        cell_value = worksheet.cell(row=row, column=id_column).value
        if cell_value:
            # Extract prefix and number (e.g., "R001" -> "R", 1)
            prefix = ''.join(c for c in str(cell_value) if c.isalpha())
            num_str = ''.join(c for c in str(cell_value) if c.isdigit())
            if num_str:
                max_id = max(max_id, int(num_str))

    if not prefix:
        # Determine prefix based on sheet name
        sheet_name = worksheet.title.lower()
        if "risk" in sheet_name:
            prefix = "R"
        elif "task" in sheet_name:
            prefix = "T"
        else:
            prefix = "ID"

    return f"{prefix}{str(max_id + 1).zfill(3)}"


def update_risk_register(workbook, risks: list, source: str) -> list:
    """Add new risks to the Risk Register sheet. Returns list of changes made."""
    ws = workbook["Risk Register"]
    changes = []
    today = datetime.now().strftime("%Y-%m-%d")

    for risk in risks:
        next_id = get_next_id(ws)
        risk_score = calculate_risk_score(risk.get("probability", "Medium"),
                                          risk.get("impact", "Medium"))

        # Find next empty row
        next_row = ws.max_row + 1

        # Columns: Risk ID, Title, Description, Category, Probability, Impact,
        # Risk Score, Status, Trend, Owner, Mitigation Plan, Linked Tasks,
        # Date Identified, Last Updated, Update History, Source, Closed Date, Resolution Notes
        ws.cell(row=next_row, column=1, value=next_id)
        ws.cell(row=next_row, column=2, value=risk.get("title", ""))
        ws.cell(row=next_row, column=3, value=risk.get("description", ""))
        ws.cell(row=next_row, column=4, value=risk.get("category", ""))
        ws.cell(row=next_row, column=5, value=risk.get("probability", "Medium"))
        ws.cell(row=next_row, column=6, value=risk.get("impact", "Medium"))
        ws.cell(row=next_row, column=7, value=risk_score)
        ws.cell(row=next_row, column=8, value="Open")
        ws.cell(row=next_row, column=9, value="New")
        ws.cell(row=next_row, column=10, value=risk.get("owner", "TBD"))
        ws.cell(row=next_row, column=11, value=risk.get("mitigation_plan", ""))
        ws.cell(row=next_row, column=12, value="")  # Linked Tasks
        ws.cell(row=next_row, column=13, value=today)  # Date Identified
        ws.cell(row=next_row, column=14, value=today)  # Last Updated
        ws.cell(row=next_row, column=15, value=f"{today}: Created from transcript")
        ws.cell(row=next_row, column=16, value=source)

        changes.append(f"Added Risk {next_id}: {risk.get('title', '')}")

    return changes


def update_tasks(workbook, tasks: list, source: str) -> list:
    """Add new tasks to the Tasks sheet. Returns list of changes made."""
    ws = workbook["Tasks"]
    changes = []
    today = datetime.now().strftime("%Y-%m-%d")

    for task in tasks:
        next_id = get_next_id(ws)

        # Find next empty row
        next_row = ws.max_row + 1

        # Columns: Task ID, Task, Owner, Due Date, Status, Linked Risk,
        # Source, Created Date, Completed Date
        ws.cell(row=next_row, column=1, value=next_id)
        ws.cell(row=next_row, column=2, value=task.get("task", ""))
        ws.cell(row=next_row, column=3, value=task.get("owner", "TBD"))
        ws.cell(row=next_row, column=4, value=task.get("due_date", ""))
        ws.cell(row=next_row, column=5, value="Open")
        ws.cell(row=next_row, column=6, value=task.get("linked_risk", ""))
        ws.cell(row=next_row, column=7, value=source)
        ws.cell(row=next_row, column=8, value=today)

        changes.append(f"Added Task {next_id}: {task.get('task', '')[:50]}...")

    return changes


def log_update(workbook, source: str, source_type: str, changes: list, raw_extract: dict):
    """Add entry to the Update Log sheet."""
    ws = workbook["Update Log"]

    # Find next empty row
    next_row = ws.max_row + 1

    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    changes_text = "\n".join(changes) if changes else "No changes made"
    raw_text = json.dumps(raw_extract, indent=2)

    # Columns: Timestamp, Source, Source Type, Changes Made, Raw Extract
    ws.cell(row=next_row, column=1, value=timestamp)
    ws.cell(row=next_row, column=2, value=source)
    ws.cell(row=next_row, column=3, value=source_type)
    ws.cell(row=next_row, column=4, value=changes_text)
    ws.cell(row=next_row, column=5, value=raw_text[:32000])  # Excel cell limit


def main():
    """Main entry point."""
    args = parse_arguments()

    # Determine script directory for relative paths
    script_dir = Path(__file__).parent.resolve()

    # Determine Risk Register path
    if args.risk_register:
        risk_register_path = Path(args.risk_register)
    else:
        risk_register_path = script_dir / "Risk_Registers" / f"Risk_Register_{args.project_code}.xlsx"

    if not risk_register_path.exists():
        print(f"Error: Risk Register not found at {risk_register_path}")
        sys.exit(1)

    # Load transcript
    print(f"Loading transcript from: {args.transcript_path}")
    try:
        transcript = load_transcript(args.transcript_path)
    except FileNotFoundError as e:
        print(f"Error: {e}")
        sys.exit(1)

    # Extract data using Claude API
    print("Extracting risks, tasks, and decisions using Claude API...")
    try:
        extracted_data = extract_data_with_claude(transcript, args.project_code)
    except anthropic.APIError as e:
        print(f"Error calling Claude API: {e}")
        sys.exit(1)
    except json.JSONDecodeError as e:
        print(f"Error parsing Claude response: {e}")
        sys.exit(1)

    print(f"  Found {len(extracted_data.get('risks', []))} risks")
    print(f"  Found {len(extracted_data.get('tasks', []))} tasks")
    print(f"  Found {len(extracted_data.get('decisions', []))} decisions")

    # Load Risk Register workbook
    print(f"Updating Risk Register: {risk_register_path}")
    workbook = openpyxl.load_workbook(risk_register_path)

    # Derive source name from transcript filename
    source_name = Path(args.transcript_path).stem

    # Update Risk Register with extracted data
    all_changes = []

    if extracted_data.get("risks"):
        risk_changes = update_risk_register(workbook, extracted_data["risks"], source_name)
        all_changes.extend(risk_changes)
        print(f"  Added {len(risk_changes)} risks to Risk Register")

    if extracted_data.get("tasks"):
        task_changes = update_tasks(workbook, extracted_data["tasks"], source_name)
        all_changes.extend(task_changes)
        print(f"  Added {len(task_changes)} tasks to Tasks sheet")

    # Log the update
    log_update(workbook, source_name, "Meeting Transcript", all_changes, extracted_data)
    print("  Logged changes to Update Log sheet")

    # Save workbook
    workbook.save(risk_register_path)
    print(f"\nSuccessfully updated {risk_register_path}")

    # Print summary
    print("\n=== Summary ===")
    for change in all_changes:
        print(f"  - {change}")

    if extracted_data.get("decisions"):
        print("\n=== Decisions Recorded ===")
        for decision in extracted_data["decisions"]:
            print(f"  - {decision.get('decision', '')}")


if __name__ == "__main__":
    main()
